""" qrep.py - query report

Take a database query file as input and outputs the query response as a formatted Excel workbook.
"""

from openpyxl.utils.exceptions import IllegalCharacterError, WorkbookAlreadySaved
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
import pyodbc
import sys


def clean_data(value):
    # Escape these chars in TEXT/str fields
    if type(value) is str: 
        value = value.replace("'", "\\")
        

def get_queryfile(filename):
    """ Read the query from the specified file. """
    try:
        with open(filename, 'r') as qry_file:
            query = qry_file.read()
            return query
    except FileNotFoundError:
        print(f"Query file '{filename}' not found.")
        return ""


def create_rows(query, cursor):
    """ Put header and data into a list of tuple rows """
    report_rows = []

    response = cursor.execute(query).fetchall()
    header = tuple([column[0] for column in cursor.description])
    
    report_rows.append(header)
    [report_rows.append(row) for row in response]
    
    return report_rows


def write_report(report_rows, xl_file):
    """ Create and format a workbook """    
    wb = Workbook(xl_file)
    ws = wb.create_sheet("Sheet1")

    # Write to worksheet.
    for row in report_rows:
        try:
            ws.append(tuple(clean_data(row)))
        except IllegalCharacterError:  # TODO: Remove illegal characters before append(), but how?
            pass                       # Illegal character currently causes blank row
        except WorkbookAlreadySaved:
            pass  # Required to avoid ws PermissionError, and doesn't affect the output

    wb.save(xl_file)


def format_worksheet(xl_file):

    wb = load_workbook(xl_file)
    ws = wb.active

    # Set cell borders, alignment,  
    for column_cells in ws.columns:
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set header properties: column width, cell color, font
    for cell in ws[1]:  # ws[1] is the first row of the sheet
        ws.column_dimensions[cell.column].width = 25
        cell.fill = PatternFill(fill_type='solid', fgColor='99CCFFFF')
        cell.font = Font(bold=True, size=14)       

    wb.save(xl_file)


def main(argv):
    
    server = 'APP-SAM-MSSQL\\APP'
    database = 'EOS'  # Transactional NPS

    if len(argv) != 3:
        print(f"\nUsage is {argv[0]} <sql qryfile.qry> <xl output.xlsx>")
        return    
    else:
        # See https://www.connectionstrings.com/ for the appropriate db connection string,
        # and OS documentation for the correct ODBC database driver.
        print(f'\nConnecting to {database}...')
        cnxn = pyodbc.connect('DRIVER={SQL Server Native Client 11.0};SERVER='+server+';DATABASE='+database+';Trusted_Connection=yes;')
        
        cursor = cnxn.cursor()     
        query = get_queryfile(argv[1])
        xl_file = argv[2]

        print(f'Creating report from {argv[1]}...')
        report_rows = create_rows(query, cursor)
        write_report(report_rows, xl_file)
        print(f'Formatting report...')
        format_worksheet(xl_file)
        print(f'Report saved to {xl_file}!\n')


if __name__ == "__main__":
    main(sys.argv)
