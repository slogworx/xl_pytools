import pytest
from pathlib import Path
from openpyxl import Workbook
from xlrd import open_workbook
from slogworx.prep_xl import (
    get_xl_json, 
    guess_type,
    fix_type,
    open_workbook,
    clean_data)


def test_get_xl_json():
    """ Create TEST_WB """
    TEST_WB = 'test_xl.xlsx'
    wb = Workbook()
    ws = wb.active

    ws['A1'], ws['B1'], ws['C1'] = 'Character', 'Weapon', 'Class'
    ws['A2'], ws['B2'], ws['C2'] = 'Fjord', 'Scimitar', 'Warlock'
    ws['A3'], ws['B3'], ws['C3'] = 'Jester', 'Lolly', 'Cleric'
    ws['A4'], ws['B4'], ws['C4'] = 'Beau', 'Fists', 'Monk'

    wb.save(TEST_WB)

    """ Make sure get_xl_json() returns a dict with the correct keys """
    xl_json_keys = ['worksheet name', 'header names', 'unique header', 'consistent type', 'col values']
    ws = open_workbook(TEST_WB).sheet_by_index(0)
    xl_json = get_xl_json(ws)
    
    assert(type(xl_json) == dict)
    
    for k in xl_json_keys:
        assert(k in list(xl_json.keys()))

    """ Delete TEST_WB """
    p = Path(TEST_WB)
    p.unlink()


def test_guess_type():
    """ When all the types are the most, including None types """
    assert(guess_type([1, "Two", "Three", 'Four', 5]) == str)
    assert(guess_type(['1', '2', 3, 4, 5]) == int)
    assert(guess_type(['1.2', 2.3, 3.4, 4.5, '5.6']) == float)

    """ Half and half number types """
    assert(guess_type(['1.2', '2.3', 3.4, 4.5]) == float)
    assert(guess_type(['One point one', 'Two point one', 3.1, 4.1]) == str)
    assert(guess_type(['1', '2', 3, 4]) == int)

    """ None can never be the most because of mostly blank columns """
    assert(guess_type([None, None, 1, 2, None]))
    assert(guess_type([None, None, None, None, None]))


def test_fix_type():
    float_types = ['1.2', '2.3', 3.4, 4.5, 5.6] 
    int_types = ['1', '2', 3, 4, 5]
    str_types = [1, 2, "three", "four", "five"]

    new_col = fix_type(float, float_types)
    for col in new_col:
        assert(type(col) is float)

    new_col = fix_type(int, int_types)
    for col in new_col:
        assert(type(col) is int)

    new_col = fix_type(str, str_types)
    for col in new_col:
        assert(type(col) is str)


def test_clean_data():
    """  Test a clean sheet """
    xl_json = { 
        'worksheet name': 'Game Sheet',
        'header names': ['Character','Class','Weapon'],
        'unique header': False,
        'consistent type': {0: True, 1: True, 2: True} ,
        'col values': [
            ['Jester', 'Fjord', 'Nott'],
            ['Cleric', 'Warlock', 'Rogue'],
            ['Lolly', 'Scimitar', 'Crossbow/Dagger']
        ]
    }
    cleaned_xl_json = clean_data(xl_json)
    assert(type(cleaned_xl_json) is dict)
    assert(cleaned_xl_json['unique header'])

    """ Test a repeated column header """
    xl_json['header names'] = ['Character', 'Level', 'Character']
    xl_json['unique header'] = False
    xl_json['consistent type'] = {0: True, 1: False, 2: True}
    xl_json['col values'] = [
        ['Jester', 'Fjord', 'Nott'],
        [9, 9, 10],
        ['Yes', 'Yes', 'Yes']
    ]
    cleaned_xl_json = clean_data(xl_json)
    assert(type(cleaned_xl_json) is dict)
    assert(cleaned_xl_json['unique header'])

    """ Test inconsistent types """
    xl_json['header names'] = ['Character', 'Level', 'Class']
    xl_json['unique header'] = False
    xl_json['consistent type'] = {0: True, 1: False, 2: True}
    xl_json['col values'] = [
        ['Jester', 'Fjord', 'Nott'],
        [9, 9, 'Ten'],
        ['Cleric', 'Warlock', 'Rogue']
    ]
    cleaned_xl_json = clean_data(xl_json)
    assert(type(cleaned_xl_json) is dict)
    assert(all(list(cleaned_xl_json['consistent type'].values())))
