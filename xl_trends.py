"""
Create an sqlite database from Excel workbook(s) 
"""
from openpyxl import load_workbook, Workbook
from slogworx.prep_xl import guess_type
from pathlib import Path
import sqlite3
import sys

def get_path(file_path):
    """ Take specified path and return a pathlib object containing all xls* files """
    p = Path(file_path)
    if p.is_dir():
        return list(p.glob('**/*.xls*'))
    else:
        print(f'{file_path} does not appear to be a valid directory.')
        return None


def dbize_data(value, header=False):
    rep_ch = [' ', '?', '#', '@', '_', '.', '-']  # Special characters that should not be in table names
    if header and type(value) is str:
        for ch in rep_ch:
            value = value.replace(ch, '')  # Eliminate special characters, too
    if header and not value is str:
        value = str(value)
    if type(value) is str: 
        value = value.replace('"', '""')  # Escape quotes in TEXT fields
        value = value.replace("'", "''")
        value = value.replace("(", "\\(")  
        value = value.replace(")", "\\)")

    return value


def read_workbook(xl_pathlib):
    """  """
    wb = load_workbook(xl_pathlib)
    
    wb_data = {}
    for sheetname in wb.sheetnames:
        data_types = [], [], [], []
        sheet = wb[sheetname]
        rows = sheet.iter_rows()
        cols = sheet.iter_cols()
        
        # Get header
        header = [dbize_data(cell.value, header=True) for cell in next(rows)]  
        # Get data
        data = [ [dbize_data(cell.value) for cell in row] for row in rows ]
        # Get types
        data_types = []
        for col in cols:
            for cell in col:
                columns = [cell.value for cell in col]
            columns.pop(0)  # Get rid of the header
            row_type = guess_type(columns)
            columns = []
            data_types.append(row_type)

        wb_data[sheetname] = {
            'wb name': sheetname,
            sheetname: dict(zip(header, data_types)),
            'data': data
        }
    
    return wb_data


def create_db(database, workbook, seperate=False):
    """ Create the database and tables. A separate table will be created for each ws if separate == False """

    cnxn = sqlite3.connect(database)
    c = cnxn.cursor()
    database_name = dbize_data(database.name, header = True)

    if seperate:
        table_name = workbook[database_name]['wb name']
    else:
        table_name = database_name

    # Construct the table CREATE string
    for sheet in workbook.keys():
        
        create_string = f'CREATE TABLE IF NOT EXISTS {database_name} (entry_id INTEGER PRIMARY KEY AUTOINCREMENT, '

        for field, data_type in workbook[sheet][sheet].items():
            create_string += f'{field} '

            if field is None:
                continue
            elif data_type is int:
                create_string += 'INT, '
            elif data_type is float:
                if 'date' in field.lower():  # Excel is not good with storing dates
                    create_string += ' DATE, '
                else:
                    create_string += ' FLOAT, '
            elif data_type is str:
                field = str(field)
                if 'date' in field.lower():
                    create_string += ' DATETIME, '
                else:
                    create_string += ' TEXT, '

        create_string = create_string[:-2] + ');'  # Get rid of trailing space and comma, then close
        
        c.execute(create_string)

    if not seperate:
        return table_name
    else:
        return workbook.keys()


def check_columns(db, wb_data):
    cnxn = sqlite3.connect(db)
    c = cnxn.cursor()

    table_name = dbize_data(db.name, header = True)
    pragma_string = f'PRAGMA table_info({table_name})'
    response = c.execute(pragma_string).fetchall()


    all_headers = False
    for sheet in wb_data.keys():
        header_names = list(wb_data[sheet][sheet].keys())
        for header_name in header_names:
            if header_name in response:
                all_headers = True
            else:
                all_headers = False
    
    return all_headers


def insert_data(database, workbook, seperate = False):
    """ Insert the worksheet data into the db table(s) """
    cnxn = sqlite3.connect(database)
    c = cnxn.cursor()
    if seperate:
        table_name = workbook['wb name']
    else:
        table_name = dbize_data(database.name, header = True)

    for sheet in workbook.keys():

        insert_string = f'INSERT INTO {table_name} ('
        for field in workbook[sheet][sheet].keys():
            insert_string += f'{field}, '
        insert_string = insert_string[:-2] + ') VALUES ('
        
        for row in workbook[sheet]['data']:
            values_string = ''
            for cell in row:
                if cell is None:
                    cell = ''
                if type(cell) is str:
                    values_string += f'"{cell}", '
                else: values_string += f'{cell}, '
            values_string = insert_string + values_string[:-2] + ');'
            c.execute(values_string)
    
    cnxn.commit()


def xl_trends(argv):
    if not len(argv) == 3:
        print(f'Usage: {argv[0]} database.db xl_path')
    else:
        database = argv[1]
        xl_path = argv[2]

        db = Path(database)
        p  = get_path(xl_path)

        if not p:
            print(f'get_path() was unable to return {xl_path}')
        elif type(p) != list: 
            print(f'{xl_path} does not appear to be a directory.')
        else:
            for xl_file in p:
                wb_data = read_workbook(xl_file)
                
                if not db.exists():
                    print(f'{db} does not exist. Creating it.')
                    create_db(db, wb_data, seperate = False)
                elif not check_columns(db, wb_data):
                    print(f'{db} exists, but is corrupted. Attempting to recreate it.')
                    db.unlink()
                    create_db(db, wb_data, seperate = False)

                insert_data(db, wb_data, seperate = False)
                

if __name__ == "__main__":
    xl_trends(sys.argv)
