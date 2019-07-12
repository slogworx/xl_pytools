"""  """
from xlrd import open_workbook
from openpyxl import Workbook, load_workbook
from pathlib import Path
from time import sleep
import sys

def get_xl_json(worksheet):
    """ Grab worksheet data and flag broken rules to fix later """
    unique_header = True
    
    # Create header
    header_names = [cell.value for cell in worksheet.row(0) if not cell.value == '']

    # Rule 1: header names must be unique
    for header in header_names:
        if header_names.count(header) != 1:
            unique_header = False

    headered_cols = len(header_names)
    # Get column values       
    col_values = [worksheet.col_values(col_num, 1) for col_num in range(headered_cols)]
    
    # Get types for every cell in the row
    col_types = [worksheet.col_types(col_num, 1) for col_num in range(headered_cols)]
    
    # Rule 2: column value types should be identical
    consistent_type = {
        n: max(col_type) == min(col_type) 
        for n, col_type in enumerate(col_types) 
    }

    # TODO: create JSON, encrypt, and save as file? Is there a reason for it?

    return { 
        'worksheet name': worksheet.name,
        'header names': header_names,
        'unique header': unique_header,
        'consistent type': consistent_type,
        'col values': col_values 
    }


def guess_type(col_values):
    """ Find the most frequent data type in col_values """
    highest_freq = 0
    ret_type = None

    values = [cell_value for cell_value in col_values]  
    type_list = [type(cell_type) for cell_type in col_values]

    values_types = zip(values, type_list)

    for cell_value, cell_type in values_types:
        if not cell_type is type(None):
            freq = lambda value_type, n: (type_list.count(value_type) + n) / len(type_list)
            
            if  freq(cell_type, 0) > 0.5:
                return cell_type
            elif freq(cell_type, 0) == 0.5 and cell_type is str:
                try:
                    if float(cell_value):
                        if freq(float, 1) > 0.5:
                            return float
                except ValueError:
                    if freq(str, 1) > 0.5:
                        return str
            else:
                curr_freq = freq(cell_type, 0)
                if curr_freq > highest_freq:
                    highest_freq = curr_freq
                    ret_type = cell_type
        
        if ret_type is None:  # This will happen if a column is completely blank, and nope.
            ret_type = str
    
    return ret_type


def numberize(cell_value, cell_type):
    ch_num = ""
    for ch in cell_value:
        if ch.isnumeric():
            ch_num += ch
    if not ch_num:
        return None
    elif cell_type is float:
        return float(ch_num)
    elif cell_type is int:
        return int(ch_num)
    else:
        return None


def fix_type(col_type, col_values):
    """ Normalize type on all columns. Keep in mind type should be dictated by guess_type(). """
    new_col = []
    for cell_value in col_values:
        curr_type = type(cell_value)
        if col_type is not str and curr_type is str: 
            new_col.append(numberize(cell_value, col_type))     
        elif col_type is float and curr_type is int:
            new_col.append(float(cell_value))
        elif col_type is str and not curr_type is str:
            new_col.append(str(cell_value))
        else:  
            new_col.append(cell_value)
    
    return new_col


def clean_data(xl_json):
    """ Enforce clean data rules from get_xl_json() """
    cleaned_col_values = []
    col_values = xl_json['col values']
    for k,v in xl_json['consistent type'].items():
        if not v:
            col_type = guess_type(col_values[k])
            cleaned_col_values.append(fix_type(col_type, col_values[k]))
            xl_json['consistent type'][k] = True
        else:
            cleaned_col_values.append(col_values[k])
    
    xl_json['col values'] = cleaned_col_values

    corrected_headers = []
    if xl_json['unique header'] is False:
        n = 1
        header_names = xl_json['header names']
        
        for name in header_names:
            matches = header_names.count(name)
            if matches > 1:
                name = f'{name}{n}'
                n += 1
            corrected_headers.append(name)

        xl_json['unique header'] = True
        xl_json['header names'] = corrected_headers
    
    return xl_json


def write_book(xl_json, wb_name):
    """  """
    wb_path = Path(wb_name)
    if wb_path.exists():
        wb = load_workbook(wb_name)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    
    ws = wb.create_sheet(xl_json['worksheet name'])
    print(f"Writing {xl_json['worksheet name']}...")
    # Write header to xlsx
    ws.append(xl_json['header names'])
    
    # Write data to xlsx
    col, row = 1, 2
    for col_cells in xl_json['col values']:
        for cell in col_cells:
            ws.cell(row, col, cell)
            row += 1
        col += 1
        row = 2

    wb.save(wb_name)


def prep_xl(argv):
    """  """
    if not len(argv) == 2:
        print(f'Usage: {argv} workbook.xls [workbook.xlsx]')
    elif not (argv[1][-5:] == '.xlsx') and not (argv[1][-4:] == '.xls'):
        print(f'{argv[1]} must be an xls or xlsx file.')
    else:
        xl_file = Path(argv[1])
        workbook = open_workbook(xl_file)
        for worksheet in workbook.sheets():
            xl_json = get_xl_json(worksheet)
            xl_json = clean_data(xl_json)
            write_book(xl_json, f'{xl_file}_PREPPED.xlsx')


if __name__ == "__main__":
    prep_xl(sys.argv)
