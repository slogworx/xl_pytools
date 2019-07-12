import sqlite3
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from time import sleep


def get_query(query_file):
    """ Open query file, read and return query string """
    p = Path(query_file)
    with p.open('r') as qf:
        query = qf.read()
    
    return query


def make_report(sqlite_data, header, report_name):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in sqlite_data:
        ws.append(row)
    
    # Set cell borders, alignment,
    for column_cells in ws.columns:
        for cell in column_cells:
            ws.row_dimensions[cell.row].height = 50
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Set header properties: column width, cell color, font
    for cell in ws[1]:  # ws[1] is the first row of the sheet
        ws.column_dimensions[cell.column_letter].width = 30
        ws.row_dimensions[cell.row].height = 25
        cell.fill = PatternFill(fill_type='solid', fgColor='99CCFFFF')
        cell.font = Font(bold=True, size=14)
    
    wb.save(report_name)


def q_reporting(argv):
    if not len(argv) == 4:
        print(f'Usage: {argv[0]} query_file.qry database.db out_file.xlsx')
        return None

    query = get_query(argv[1])
    db = Path(argv[2])
    out_file = Path(argv[3])
    
    cnxn = sqlite3.connect(db)
    c = cnxn.cursor()

    response = c.execute(query)
    col_fields = response.description
    header =  []
    for col in col_fields:
        header.append(col[0])    
    make_report(response.fetchall(), tuple(header), out_file)


if __name__ == "__main__":
    q_reporting(sys.argv)